import selenium 
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import time 

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service = service, options = option)
driver.maximize_window()
url = "https://www.1a.lv/"
driver.get(url)
time.sleep(2)

accept_cookie = driver.find_element(By.CLASS_NAME, "c-button")
accept_cookie.click()
time.sleep(2)

search_tab = driver.find_element(By.ID, "q")
search_tab.click()
search_tab.send_keys("RTX 4060")
time.sleep(2)

search_submit = driver.find_element(By.CLASS_NAME, "main-search-submit")
search_submit.click()
time.sleep(3)

GPU_section = driver.find_element(By.CSS_SELECTOR, '[alt="Video kartes"]')
GPU_section.click()
time.sleep(2)

price_to = driver.find_element(By.ID, "attr-to-priceLoyalty")
time.sleep(2)
driver.execute_script("window.scrollTo(0,500)")
time.sleep(2)
price_to.click()
time.sleep(2)
price_to.send_keys(Keys.CONTROL + "a")
time.sleep(2)
price_to.send_keys(Keys.DELETE)
price_to.send_keys("370")
price_to.send_keys(Keys.ENTER)
time.sleep(2)

GPU_titles_list = []
GPU_prices_list = []

GPU_title_element = driver.find_elements(By.CLASS_NAME, "ks-new-product-name")
GPU_price_element = driver.find_elements(By.CLASS_NAME, "ks-item-price")
GPU_PSU_element = driver.find_elements(By.CLASS_NAME, "ks-new-product-attributes")

GPU_titles_list = []
GPU_prices_list = []

for i in range(len(GPU_title_element)):
    title = GPU_title_element[i]
    price = GPU_price_element[i]
    psu_power = GPU_PSU_element[i]

    GPU_title_text = title.text
    GPU_price_text = price.text

    if "RTX™ 4060" in GPU_title_text and ("550 W" in psu_power.text):
        GPU_titles_list.append(GPU_title_text)
        GPU_prices_list.append(GPU_price_text)

wb = Workbook()
ws = wb.active
max_row = ws.max_row

ws['A1'] = "GPU NAME"
ws['B1'] = "GPU PRICE"

ws['A1'].font = Font(bold = True)
ws['B1'].font = Font(bold = True)

ws.column_dimensions['A'].width = 70
ws.column_dimensions['B'].width = 15

for col in ['A']:
    ws[col + '1'].alignment = Alignment(horizontal='center')

for col in ['B']:
    ws[col + '1'].alignment = Alignment(horizontal='center')

for i in range(len(GPU_titles_list)):
    row = max_row + 2 + i
    ws['A' + str(row)].value = GPU_titles_list[i]
    ws['B' + str(row)].value = GPU_prices_list[i]

driver.quit()
wb.save('result.xlsx')
wb.close()
